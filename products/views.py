"""products/views.py"""
import io
import logging

import openpyxl
from django.db.models import Q
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, generics, status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import IsSeller, IsSellerOrAdmin
from core.pagination import StandardResultsPagination
from .filters import ProductFilter
from .models import Product, Category
from .serializers import (
    CategorySerializer,
    ProductSerializer,
    ProductCreateSerializer,
    ProductBulkRowSerializer,
    ProductCompareSerializer,
)

logger = logging.getLogger(__name__)


class CategoryListView(generics.ListAPIView):
    """Public list of all product categories."""
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [AllowAny]


class ProductListView(generics.ListAPIView):
    """Public paginated product listing with filtering & search."""
    serializer_class = ProductSerializer
    permission_classes = [AllowAny]
    pagination_class = StandardResultsPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProductFilter
    search_fields = ['name', 'description', 'SKU', 'seller__name', 'category__name']
    ordering_fields = ['price', 'created_at', 'stock_qty', 'avg_rating']
    ordering = ['-created_at']

    def get_queryset(self):
        from django.db.models import Avg
        return (
            Product.objects.filter(is_active=True)
            .annotate(avg_rating=Avg('reviews__stars'))
            .select_related('seller', 'seller__seller_profile', 'category')
            .prefetch_related('images', 'reviews')
        )


class ProductDetailView(generics.RetrieveAPIView):
    """Single product detail (public)."""
    serializer_class = ProductSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return (
            Product.objects.filter(is_active=True)
            .select_related('seller', 'seller__seller_profile', 'category')
            .prefetch_related('images', 'reviews')
        )


class ProductCreateView(generics.CreateAPIView):
    """Create a product (sellers only)."""
    serializer_class = ProductCreateSerializer
    permission_classes = [IsSeller]
    parser_classes = [MultiPartParser, FormParser, JSONParser]


class ProductUpdateView(generics.UpdateAPIView):
    """Update a product (owner seller or admin only)."""
    serializer_class = ProductCreateSerializer
    permission_classes = [IsSeller]
    http_method_names = ['put', 'patch']

    def get_queryset(self):
        return Product.objects.filter(seller=self.request.user)


class ProductSearchView(generics.ListAPIView):
    """Full text search across products."""
    serializer_class = ProductSerializer
    permission_classes = [AllowAny]
    pagination_class = StandardResultsPagination

    def get_queryset(self):
        q = self.request.query_params.get('q', '').strip()
        if not q:
            return Product.objects.none()
        return (
            Product.objects.filter(
                Q(name__icontains=q)
                | Q(description__icontains=q)
                | Q(SKU__icontains=q)
                | Q(seller__name__icontains=q)
                | Q(category__name__icontains=q),
                is_active=True,
            )
            .select_related('seller', 'seller__seller_profile', 'category')
            .prefetch_related('images', 'reviews')
            .order_by('-created_at')
        )


class ProductCompareView(APIView):
    """
    Price Comparison Engine.
    Returns all listings of same-named products with lowest price identified.
    """
    permission_classes = [AllowAny]

    def get(self, request, pk):
        try:
            base_product = Product.objects.select_related(
                'seller', 'seller__seller_profile'
            ).get(pk=pk, is_active=True)
        except Product.DoesNotExist:
            return Response(
                {'status': 'error', 'message': 'Product not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Find comparable products by name similarity
        similar = (
            Product.objects.filter(
                name__icontains=base_product.name.split()[0],
                is_active=True,
            )
            .select_related('seller', 'seller__seller_profile')
            .order_by('price')
        )

        if not similar.exists():
            similar = Product.objects.filter(pk=pk).select_related(
                'seller', 'seller__seller_profile'
            )

        lowest_price = similar.first().price if similar.exists() else base_product.price

        comparison = []
        for product in similar:
            try:
                seller_rating = product.seller.seller_profile.rating_avg
            except Exception:
                seller_rating = 0

            comparison.append({
                'product_id': str(product.id),
                'product_name': product.name,
                'seller_name': product.seller.name,
                'seller_rating': seller_rating,
                'price': product.price,
                'stock_qty': product.stock_qty,
                'delivery_days': product.delivery_days,
                'is_lowest_price': product.price == lowest_price,
                'price_difference': product.price - lowest_price,
            })

        serializer = ProductCompareSerializer(comparison, many=True)
        return Response({
            'status': 'success',
            'base_product': ProductSerializer(base_product).data,
            'comparison_count': len(comparison),
            'lowest_price': str(lowest_price),
            'comparison': serializer.data,
        })


class ProductBulkUploadView(APIView):
    """
    Bulk product upload via Excel (.xlsx).
    Validates every row and returns a detailed error report.
    """
    permission_classes = [IsSeller]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'status': 'error', 'message': 'No file provided.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not file.name.endswith('.xlsx'):
            return Response(
                {'status': 'error', 'message': 'Only .xlsx files are accepted.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            return Response(
                {'status': 'error', 'message': f'Could not parse file: {e}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        expected = ['name', 'description', 'price', 'stock_qty', 'SKU', 'category']
        missing = [h for h in expected if h not in headers]
        if missing:
            return Response(
                {'status': 'error', 'message': f'Missing columns: {missing}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        col_map = {v: i for i, v in enumerate(headers)}
        created, errors = [], []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            row_data = {
                'name': row[col_map['name']],
                'description': row[col_map.get('description', -1)] or '',
                'price': row[col_map['price']],
                'stock_qty': row[col_map['stock_qty']],
                'SKU': row[col_map['SKU']],
                'category': row[col_map.get('category', -1)] or '',
            }
            serializer = ProductBulkRowSerializer(data=row_data)
            if serializer.is_valid():
                data = serializer.validated_data
                category_name = data.pop('category', '')
                category = None
                if category_name:
                    category, _ = Category.objects.get_or_create(
                        name=category_name,
                        defaults={'slug': category_name.lower().replace(' ', '-')}
                    )
                try:
                    Product.objects.create(
                        seller=request.user,
                        category=category,
                        **data,
                    )
                    created.append(row_data['SKU'])
                except Exception as e:
                    errors.append({'row': row_num, 'SKU': row_data.get('SKU'), 'error': str(e)})
            else:
                errors.append({'row': row_num, 'data': row_data, 'errors': serializer.errors})

        return Response({
            'status': 'success',
            'created_count': len(created),
            'error_count': len(errors),
            'created_skus': created,
            'errors': errors,
        }, status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED)
